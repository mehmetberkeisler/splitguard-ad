#!/usr/bin/env python3
"""Build the OASIS-1 SplitGuard-AD artifacts.

This script avoids full archive extraction. It selectively extracts only the
processed masked T88 Analyze pairs needed for the OASIS-1 replication study,
then creates volume and slice manifests, a subject-level leakage graph, a
component-safe split, and audit reports.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import random
import re
import shutil
import tarfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import nibabel as nib
import numpy as np
from openpyxl import load_workbook
from PIL import Image


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OASIS_ROOT = PROJECT_ROOT / "oasis1"
TAR_DIR = OASIS_ROOT / "tar"
METADATA_DIR = OASIS_ROOT / "metadata"
SELECTED_DIR = OASIS_ROOT / "selected"
VOLUME_DIR = SELECTED_DIR / "processed_t88_masked"
SLICE_DIR = SELECTED_DIR / "slices"

MANIFEST_DIR = PROJECT_ROOT / "data" / "manifests"
SPLIT_DIR = PROJECT_ROOT / "data" / "splits"
AUDIT_DIR = PROJECT_ROOT / "reports" / "audits"

INVENTORY_CSV = MANIFEST_DIR / "oasis1_archive_inventory.csv"
VOLUME_MANIFEST_CSV = MANIFEST_DIR / "oasis1_manifest.csv"
SLICE_MANIFEST_CSV = MANIFEST_DIR / "oasis1_slices_manifest.csv"
COMPONENT_CSV = MANIFEST_DIR / "oasis1_leakage_components.csv"
SPLIT_CSV = SPLIT_DIR / "oasis1_splitguard_seed42.csv"

INVENTORY_AUDIT_MD = AUDIT_DIR / "oasis1_inventory_audit.md"
MANIFEST_AUDIT_MD = AUDIT_DIR / "oasis1_manifest_audit.md"
SPLIT_AUDIT_MD = AUDIT_DIR / "oasis1_splitguard_seed42_audit.md"
SUMMARY_JSON = AUDIT_DIR / "oasis1_pipeline_summary.json"

SOURCE_DATASET = "oasis1_cross_sectional"
SPLIT_ORDER = ["train", "val", "test"]
DEFAULT_RATIOS = {"train": 0.70, "val": 0.15, "test": 0.15}
SLICE_AXIS = 1
SLICE_OFFSETS = [-20, -10, 0, 10, 20]


@dataclass(frozen=True)
class MetadataRow:
    session_id: str
    subject_id: str
    sex: str
    hand: str
    age: str
    educ: str
    ses: str
    mmse: str
    cdr: str
    etiv: str
    nwbv: str
    asf: str
    delay: str
    is_reliability_session: bool
    binary_label: str
    clinical_label: str
    label_confidence: str


def stable_token(text: str, length: int = 12) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()[:length]


def stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def session_to_subject(session_id: str) -> str:
    match = re.match(r"^(OAS1_\d{4})_MR\d+$", session_id)
    if not match:
        raise ValueError(f"Unsupported OASIS session ID: {session_id}")
    return match.group(1)


def label_from_cdr(value: object) -> tuple[str, str, str]:
    if value is None or stringify(value).upper() == "N/A":
        return "label_missing", "CDR_missing", "missing_cdr_excluded_from_primary"
    cdr = float(value)
    if cdr == 0.0:
        return "NonDemented", "CDR_0", "clinical_cdr"
    if cdr > 0.0:
        return "Demented", f"CDR_{stringify(value)}", "clinical_cdr"
    raise ValueError(f"Unsupported CDR value: {value}")


def metadata_workbook_path() -> Path:
    candidates = [
        path
        for path in METADATA_DIR.glob("*.xlsx")
        if "reliability" not in path.name.lower()
    ]
    if len(candidates) != 1:
        raise FileNotFoundError(
            f"Expected exactly one OASIS-1 clinical metadata workbook in {METADATA_DIR}, "
            f"found {len(candidates)}."
        )
    return candidates[0]


def reliability_workbook_path() -> Path | None:
    candidates = [
        path
        for path in METADATA_DIR.glob("*.xlsx")
        if "reliability" in path.name.lower()
    ]
    return candidates[0] if candidates else None


def read_reliability_ids() -> set[str]:
    path = reliability_workbook_path()
    if path is None:
        return set()
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    ids: set[str] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            ids.add(str(row[0]))
    return ids


def read_metadata() -> list[MetadataRow]:
    reliability_ids = read_reliability_ids()
    workbook = load_workbook(metadata_workbook_path(), read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows: list[MetadataRow] = []

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not values or not values[0]:
            continue
        raw = dict(zip(headers, values))
        session_id = str(raw["ID"])
        binary_label, clinical_label, label_confidence = label_from_cdr(raw.get("CDR"))
        rows.append(
            MetadataRow(
                session_id=session_id,
                subject_id=session_to_subject(session_id),
                sex=stringify(raw.get("M/F")),
                hand=stringify(raw.get("Hand")),
                age=stringify(raw.get("Age")),
                educ=stringify(raw.get("Educ")),
                ses=stringify(raw.get("SES")),
                mmse=stringify(raw.get("MMSE")),
                cdr=stringify(raw.get("CDR")),
                etiv=stringify(raw.get("eTIV")),
                nwbv=stringify(raw.get("nWBV")),
                asf=stringify(raw.get("ASF")),
                delay=stringify(raw.get("Delay")),
                is_reliability_session=session_id.endswith("_MR2") or session_id in reliability_ids,
                binary_label=binary_label,
                clinical_label=clinical_label,
                label_confidence=label_confidence,
            )
        )
    return rows


def archive_for_session(session_id: str) -> tuple[int, Path]:
    match = re.match(r"^OAS1_(\d{4})_(MR\d+)$", session_id)
    if not match:
        raise ValueError(f"Unsupported OASIS session ID: {session_id}")
    number = int(match.group(1))
    visit = match.group(2)
    ranges = [
        (1, 1, 42),
        (2, 43, 80),
        (3, 80, 115),
        (4, 116, 150),
        (5, 151, 191),
        (6, 192, 231),
        (7, 232, 272),
        (8, 273, 309),
        (9, 310, 348),
        (10, 349, 382),
        (11, 383, 419),
        (12, 420, 457),
    ]
    if number == 80 and visit == "MR2":
        disc = 3
    else:
        disc = next((d for d, low, high in ranges if low <= number <= high), None)
    if disc is None:
        raise ValueError(f"Could not map session {session_id} to an OASIS disc")
    archive = TAR_DIR / f"oasis_cross-sectional_disc{disc}.tar"
    return disc, archive


def expected_member_candidates(session_id: str, disc: int) -> dict[str, list[str]]:
    """Return accepted source member names for the selected processed volume.

    Most OASIS-1 sessions use the n4 processed MPRAGE name, but a subset uses
    other n-values (for example n3, n5, or n6). All accepted source variants
    are normalized into the same selected output name.
    """
    candidates: dict[str, list[str]] = {"hdr": [], "img": []}
    for version in [f"n{index}" for index in range(1, 9)]:
        base = (
            f"disc{disc}/{session_id}/PROCESSED/MPRAGE/T88_111/"
            f"{session_id}_mpr_{version}_anon_111_t88_masked_gfc"
        )
        candidates["hdr"].append(f"{base}.hdr")
        candidates["img"].append(f"{base}.img")
    return candidates


def volume_paths(session_id: str) -> tuple[Path, Path]:
    base = VOLUME_DIR / session_id / f"{session_id}_mpr_n4_anon_111_t88_masked_gfc"
    return base.with_suffix(".hdr"), base.with_suffix(".img")


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def markdown_table(headers: list[str], rows: Iterable[Iterable[object]]) -> str:
    out = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        out.append("| " + " | ".join(str(value) for value in row) + " |")
    return "\n".join(out)


def build_inventory(metadata_rows: list[MetadataRow]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for meta in metadata_rows:
        disc, archive = archive_for_session(meta.session_id)
        hdr_path, img_path = volume_paths(meta.session_id)
        members = expected_member_candidates(meta.session_id, disc)
        rows.append(
            {
                "session_id": meta.session_id,
                "subject_id": meta.subject_id,
                "disc": disc,
                "archive_path": str(archive),
                "archive_exists": archive.exists(),
                "expected_hdr_member": ";".join(members["hdr"]),
                "expected_img_member": ";".join(members["img"]),
                "selected_hdr_path": str(hdr_path),
                "selected_img_path": str(img_path),
                "selected_hdr_exists": hdr_path.exists(),
                "selected_img_exists": img_path.exists(),
                "is_reliability_session": meta.is_reliability_session,
                "binary_label": meta.binary_label,
                "clinical_label": meta.clinical_label,
            }
        )
    return rows


def write_inventory_audit(inventory_rows: list[dict[str, object]]) -> None:
    archive_counts = Counter(row["disc"] for row in inventory_rows)
    archives_missing = sorted({row["archive_path"] for row in inventory_rows if not row["archive_exists"]})
    extracted_pairs = sum(
        bool(row["selected_hdr_exists"]) and bool(row["selected_img_exists"])
        for row in inventory_rows
    )
    lines = [
        "# OASIS-1 Archive Inventory Audit",
        "",
        "## Summary",
        "",
        f"- Inventory CSV: `{INVENTORY_CSV.relative_to(PROJECT_ROOT)}`",
        f"- Clinical/session rows: **{len(inventory_rows)}**",
        f"- Unique subjects: **{len({row['subject_id'] for row in inventory_rows})}**",
        f"- Archives expected: **12**",
        f"- Archives missing: **{len(archives_missing)}**",
        f"- Selected processed volume pairs currently extracted: **{extracted_pairs}**",
        "",
        "## Sessions By Archive",
        "",
        markdown_table(
            ["Disc", "Sessions"],
            [[disc, archive_counts[disc]] for disc in sorted(archive_counts)],
        ),
        "",
        "## QC Decision",
        "",
    ]
    if archives_missing:
        lines.append("**NO-GO: one or more OASIS-1 raw archives are missing.**")
        lines.append("")
        lines.append("Missing archives:")
        lines.extend(f"- `{path}`" for path in archives_missing)
    else:
        lines.append(
            "**GO for selective extraction: all 12 OASIS-1 raw archives referenced "
            "by the metadata are present.**"
        )
    INVENTORY_AUDIT_MD.parent.mkdir(parents=True, exist_ok=True)
    INVENTORY_AUDIT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def safe_copy_member(tar: tarfile.TarFile, member: tarfile.TarInfo, destination: Path) -> None:
    source = tar.extractfile(member)
    if source is None:
        raise ValueError(f"Could not extract member: {member.name}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    with source, destination.open("wb") as handle:
        shutil.copyfileobj(source, handle)


def extract_selected_volumes(metadata_rows: list[MetadataRow], force: bool = False) -> dict[str, object]:
    desired_by_archive: dict[Path, dict[str, Path]] = defaultdict(dict)
    expected_destinations: set[Path] = set()
    for meta in metadata_rows:
        disc, archive = archive_for_session(meta.session_id)
        hdr_path, img_path = volume_paths(meta.session_id)
        members = expected_member_candidates(meta.session_id, disc)
        expected_destinations.update({hdr_path, img_path})
        for member_name in members["hdr"]:
            desired_by_archive[archive][member_name] = hdr_path
        for member_name in members["img"]:
            desired_by_archive[archive][member_name] = img_path

    found_destinations: set[Path] = set()
    found_members: set[str] = set()
    for archive in sorted(desired_by_archive, key=lambda path: int(re.search(r"disc(\d+)", path.name).group(1))):
        desired = desired_by_archive[archive]
        if not archive.exists():
            raise FileNotFoundError(f"Missing OASIS archive: {archive}")
        unique_destinations = {str(path) for path in desired.values()}
        print(f"Scanning {archive.name} for {len(unique_destinations)} selected files...")
        with tarfile.open(archive) as tar:
            for member in tar:
                destination = desired.get(member.name)
                if destination is None:
                    continue
                found_members.add(member.name)
                found_destinations.add(destination)
                if destination.exists() and not force:
                    continue
                safe_copy_member(tar, member, destination)

    missing = sorted(str(path) for path in expected_destinations - found_destinations)
    return {
        "expected_selected_files": len(expected_destinations),
        "found_selected_files": len(found_destinations),
        "found_source_members": len(found_members),
        "missing_selected_files": missing,
    }


def volume_record(meta: MetadataRow) -> dict[str, object]:
    disc, archive = archive_for_session(meta.session_id)
    hdr_path, img_path = volume_paths(meta.session_id)
    volume_id = f"vol_{stable_token(meta.session_id)}"
    paths_exist = hdr_path.exists() and img_path.exists()
    image_shape = ""
    voxel_sizes = ""
    if paths_exist:
        image = nib.load(str(hdr_path))
        image_shape = "x".join(str(value) for value in image.shape)
        voxel_sizes = "x".join(f"{value:g}" for value in image.header.get_zooms()[: len(image.shape)])
    return {
        "volume_id": volume_id,
        "session_id": meta.session_id,
        "subject_id": meta.subject_id,
        "scan_id": meta.session_id,
        "path": str(hdr_path.resolve()),
        "img_path": str(img_path.resolve()),
        "relative_path": str(hdr_path.relative_to(PROJECT_ROOT)),
        "source_dataset": SOURCE_DATASET,
        "source_archive": archive.name,
        "source_disc": disc,
        "raw_class_label": meta.clinical_label,
        "binary_label": meta.binary_label,
        "clinical_label": meta.clinical_label,
        "label_confidence": meta.label_confidence,
        "sex": meta.sex,
        "hand": meta.hand,
        "age": meta.age,
        "educ": meta.educ,
        "ses": meta.ses,
        "mmse": meta.mmse,
        "cdr": meta.cdr,
        "etiv": meta.etiv,
        "nwbv": meta.nwbv,
        "asf": meta.asf,
        "delay": meta.delay,
        "is_reliability_session": meta.is_reliability_session,
        "hdr_exists": hdr_path.exists(),
        "img_exists": img_path.exists(),
        "image_shape": image_shape,
        "voxel_sizes": voxel_sizes,
        "preprocessing_version": "oasis1_processed_mprage_t88_111_masked_gfc",
    }


def build_volume_manifest(metadata_rows: list[MetadataRow]) -> list[dict[str, object]]:
    rows = [volume_record(meta) for meta in metadata_rows]
    fieldnames = list(rows[0].keys()) if rows else []
    write_csv(VOLUME_MANIFEST_CSV, rows, fieldnames)
    return rows


def normalize_slice(slice_2d: np.ndarray) -> Image.Image:
    data = np.asarray(slice_2d, dtype=np.float32)
    data = np.nan_to_num(data, nan=0.0, posinf=0.0, neginf=0.0)
    nonzero = data[data > 0]
    if nonzero.size:
        low, high = np.percentile(nonzero, [1, 99])
    else:
        low, high = float(np.min(data)), float(np.max(data))
    if high <= low:
        high = low + 1.0
    data = np.clip((data - low) / (high - low), 0.0, 1.0)
    data = (data * 255).astype(np.uint8)
    data = np.rot90(data)
    return Image.fromarray(data, mode="L")


def generate_slices(volume_rows: list[dict[str, object]], force: bool = False) -> list[dict[str, object]]:
    slice_rows: list[dict[str, object]] = []
    for volume in volume_rows:
        hdr_path = Path(str(volume["path"]))
        if not hdr_path.exists():
            continue
        image = nib.load(str(hdr_path))
        data = np.squeeze(image.get_fdata(dtype=np.float32))
        if data.ndim != 3:
            raise ValueError(f"Expected a 3D volume after squeezing {hdr_path}, got shape {data.shape}")
        center = data.shape[SLICE_AXIS] // 2
        slice_pairs = [
            (offset, center + offset)
            for offset in SLICE_OFFSETS
            if 0 <= center + offset < data.shape[SLICE_AXIS]
        ]
        for offset, index in slice_pairs:
            slice_id = f"slice_{stable_token(str(volume['session_id']) + '_' + str(index))}"
            output_dir = SLICE_DIR / str(volume["session_id"])
            output_path = output_dir / f"{volume['session_id']}_cor_{index:03d}.png"
            if force or not output_path.exists():
                slice_image = normalize_slice(np.take(data, index, axis=SLICE_AXIS))
                output_dir.mkdir(parents=True, exist_ok=True)
                slice_image.save(output_path)
            slice_rows.append(
                {
                    "image_id": slice_id,
                    "volume_id": volume["volume_id"],
                    "session_id": volume["session_id"],
                    "subject_id": volume["subject_id"],
                    "scan_id": volume["scan_id"],
                    "path": str(output_path.resolve()),
                    "relative_path": str(output_path.relative_to(PROJECT_ROOT)),
                    "source_dataset": volume["source_dataset"],
                    "source_archive": volume["source_archive"],
                    "source_disc": volume["source_disc"],
                    "raw_class_label": volume["raw_class_label"],
                    "binary_label": volume["binary_label"],
                    "clinical_label": volume["clinical_label"],
                    "label_confidence": volume["label_confidence"],
                    "sex": volume["sex"],
                    "age": volume["age"],
                    "educ": volume["educ"],
                    "ses": volume["ses"],
                    "mmse": volume["mmse"],
                    "cdr": volume["cdr"],
                    "etiv": volume["etiv"],
                    "nwbv": volume["nwbv"],
                    "asf": volume["asf"],
                    "delay": volume["delay"],
                    "is_reliability_session": volume["is_reliability_session"],
                    "slice_axis": SLICE_AXIS,
                    "slice_index": index,
                    "slice_offset_from_center": offset,
                    "derived_from": volume["relative_path"],
                    "preprocessing_version": "oasis1_t88_masked_coronal_center5_png_v1",
                }
            )
    if slice_rows:
        write_csv(SLICE_MANIFEST_CSV, slice_rows, list(slice_rows[0].keys()))
    return slice_rows


def build_components(slice_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    by_subject: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in slice_rows:
        by_subject[str(row["subject_id"])].append(row)

    output_rows: list[dict[str, object]] = []
    for subject_id, rows in sorted(by_subject.items()):
        component_id = f"oasis1_comp_{stable_token(subject_id)}"
        non_missing_labels = sorted(
            {str(row["binary_label"]) for row in rows if row["binary_label"] != "label_missing"}
        )
        if len(non_missing_labels) == 1:
            component_label = non_missing_labels[0]
        elif len(non_missing_labels) == 0:
            component_label = "label_missing"
        else:
            component_label = "mixed_label"
        for row in rows:
            out = dict(row)
            out.update(
                {
                    "component_id": component_id,
                    "component_size": len(rows),
                    "component_primary_reason": "same_subject",
                    "component_binary_label": component_label,
                    "subject_id_confidence": "true_oasis_subject_id",
                    "subject_parse_status": "metadata_subject_id",
                }
            )
            output_rows.append(out)

    if output_rows:
        write_csv(COMPONENT_CSV, output_rows, list(output_rows[0].keys()))
    return output_rows


def class_targets(total: int, ratios: dict[str, float]) -> dict[str, int]:
    train = round(total * ratios["train"])
    val = round(total * ratios["val"])
    test = total - train - val
    return {"train": train, "val": val, "test": test}


def choose_subset_by_size(components: list[dict[str, object]], target: int) -> set[str]:
    if target <= 0:
        return set()
    dp: dict[int, tuple[str, ...]] = {0: ()}
    for component in components:
        component_id = str(component["component_id"])
        size = int(component["n_images"])
        additions: dict[int, tuple[str, ...]] = {}
        for current_sum, chosen in dp.items():
            new_sum = current_sum + size
            if new_sum not in dp and new_sum not in additions:
                additions[new_sum] = (*chosen, component_id)
        dp.update(additions)
    best_sum = min(dp, key=lambda value: (abs(value - target), value > target, value))
    return set(dp[best_sum])


def make_split(component_rows: list[dict[str, object]], seed: int = 42) -> tuple[list[dict[str, object]], dict]:
    labeled = [row for row in component_rows if row["binary_label"] in {"NonDemented", "Demented"}]
    by_component: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in labeled:
        by_component[str(row["component_id"])].append(row)

    components: list[dict[str, object]] = []
    excluded_components: list[dict[str, object]] = []
    for component_id, rows in by_component.items():
        labels = sorted({str(row["binary_label"]) for row in rows})
        if len(labels) != 1:
            excluded_components.append({"component_id": component_id, "labels": labels})
            continue
        components.append(
            {
                "component_id": component_id,
                "n_images": len(rows),
                "binary_label": labels[0],
                "subject_id": rows[0]["subject_id"],
            }
        )

    rng = random.Random(seed)
    assignments: dict[str, str] = {}
    target_rows: list[dict[str, object]] = []
    for label in ["NonDemented", "Demented"]:
        label_components = [component for component in components if component["binary_label"] == label]
        rng.shuffle(label_components)
        label_components.sort(key=lambda component: int(component["n_images"]), reverse=True)
        total_images = sum(int(component["n_images"]) for component in label_components)
        targets = class_targets(total_images, DEFAULT_RATIOS)
        val_components = choose_subset_by_size(label_components, targets["val"])
        remaining = [component for component in label_components if component["component_id"] not in val_components]
        test_components = choose_subset_by_size(remaining, targets["test"])

        achieved = Counter()
        component_counts = Counter()
        for component in label_components:
            component_id = str(component["component_id"])
            if component_id in val_components:
                split = "val"
            elif component_id in test_components:
                split = "test"
            else:
                split = "train"
            assignments[component_id] = split
            achieved[split] += int(component["n_images"])
            component_counts[split] += 1

        for split in SPLIT_ORDER:
            target_rows.append(
                {
                    "binary_label": label,
                    "split": split,
                    "target_images": targets[split],
                    "achieved_images": achieved[split],
                    "component_count": component_counts[split],
                }
            )

    split_rows: list[dict[str, object]] = []
    for row in labeled:
        component_id = str(row["component_id"])
        if component_id not in assignments:
            continue
        out = dict(row)
        out.update(
            {
                "split": assignments[component_id],
                "split_policy": "oasis1_subject_component_safe_binary_cdr_v1",
                "split_seed": seed,
            }
        )
        split_rows.append(out)
    split_rows.sort(key=lambda row: (str(row["split"]), str(row["subject_id"]), str(row["relative_path"])))
    if split_rows:
        write_csv(SPLIT_CSV, split_rows, list(split_rows[0].keys()))

    summary = summarize_split(split_rows, target_rows, excluded_components)
    return split_rows, summary


def summarize_split(
    split_rows: list[dict[str, object]],
    target_rows: list[dict[str, object]],
    excluded_components: list[dict[str, object]],
) -> dict:
    images_by_split = Counter(row["split"] for row in split_rows)
    labels_by_split: dict[str, Counter] = {split: Counter() for split in SPLIT_ORDER}
    components_by_split: dict[str, set[str]] = {split: set() for split in SPLIT_ORDER}
    subjects_by_split: dict[str, set[str]] = {split: set() for split in SPLIT_ORDER}
    for row in split_rows:
        split = str(row["split"])
        labels_by_split[split][str(row["binary_label"])] += 1
        components_by_split[split].add(str(row["component_id"]))
        subjects_by_split[split].add(str(row["subject_id"]))

    component_overlap = {}
    subject_overlap = {}
    for left_index, left in enumerate(SPLIT_ORDER):
        for right in SPLIT_ORDER[left_index + 1 :]:
            component_overlap[f"{left}_vs_{right}"] = sorted(components_by_split[left] & components_by_split[right])
            subject_overlap[f"{left}_vs_{right}"] = sorted(subjects_by_split[left] & subjects_by_split[right])

    return {
        "split_manifest": str(SPLIT_CSV),
        "total_images": len(split_rows),
        "images_by_split": dict(images_by_split),
        "labels_by_split": {split: dict(labels_by_split[split]) for split in SPLIT_ORDER},
        "components_by_split": {split: len(components_by_split[split]) for split in SPLIT_ORDER},
        "subjects_by_split": {split: len(subjects_by_split[split]) for split in SPLIT_ORDER},
        "target_rows": target_rows,
        "excluded_components": excluded_components,
        "component_overlap": component_overlap,
        "subject_overlap": subject_overlap,
        "overlap_check_passed": all(not values for values in component_overlap.values())
        and all(not values for values in subject_overlap.values()),
    }


def write_manifest_audit(volume_rows: list[dict[str, object]], slice_rows: list[dict[str, object]]) -> None:
    sessions = {row["session_id"] for row in volume_rows}
    subjects = {row["subject_id"] for row in volume_rows}
    label_counts = Counter(row["binary_label"] for row in volume_rows)
    cdr_counts = Counter(row["cdr"] if row["cdr"] else "missing" for row in volume_rows)
    extracted_pairs = sum(bool(row["hdr_exists"]) and bool(row["img_exists"]) for row in volume_rows)
    slice_label_counts = Counter(row["binary_label"] for row in slice_rows)
    lines = [
        "# OASIS-1 Manifest Audit",
        "",
        "## Summary",
        "",
        f"- Volume manifest: `{VOLUME_MANIFEST_CSV.relative_to(PROJECT_ROOT)}`",
        f"- Slice manifest: `{SLICE_MANIFEST_CSV.relative_to(PROJECT_ROOT)}`",
        f"- Sessions in metadata: **{len(sessions)}**",
        f"- Unique subjects: **{len(subjects)}**",
        f"- Extracted processed volume pairs: **{extracted_pairs}**",
        f"- Derived slices: **{len(slice_rows)}**",
        f"- Slice policy: coronal axis `{SLICE_AXIS}`, offsets `{SLICE_OFFSETS}`",
        "",
        "## Volume Labels",
        "",
        markdown_table(["Label", "Sessions"], sorted(label_counts.items())),
        "",
        "## CDR Distribution",
        "",
        markdown_table(["CDR", "Sessions"], sorted(cdr_counts.items(), key=lambda item: str(item[0]))),
        "",
        "## Slice Labels",
        "",
        markdown_table(["Label", "Slices"], sorted(slice_label_counts.items())),
        "",
        "## Clinical Claim Level",
        "",
        "OASIS-1 is used as a metadata-rich replication cohort. Primary training excludes sessions with missing CDR labels, but those sessions remain visible in the manifest for auditability.",
        "",
        "## QC Decision",
        "",
    ]
    if extracted_pairs == len(volume_rows) and len(slice_rows) > 0:
        lines.append("**GO for leakage graph and subject/session-safe split generation.**")
    else:
        lines.append("**NO-GO: one or more expected processed volume pairs or derived slices are missing.**")
    MANIFEST_AUDIT_MD.parent.mkdir(parents=True, exist_ok=True)
    MANIFEST_AUDIT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_split_audit(summary: dict) -> None:
    lines = [
        "# OASIS-1 SplitGuard Seed 42 Split Audit",
        "",
        "## Summary",
        "",
        f"- Split manifest: `{SPLIT_CSV.relative_to(PROJECT_ROOT)}`",
        "- Split policy: `oasis1_subject_component_safe_binary_cdr_v1`",
        "- Component rule: same OASIS subject, including MR1/MR2 reliability sessions",
        "- Seed: **42**",
        f"- Total labeled slices: **{summary['total_images']}**",
        f"- Overlap check passed: **{summary['overlap_check_passed']}**",
        "",
        "## Images, Components, And Subjects By Split",
        "",
        markdown_table(
            ["Split", "Images", "Components", "Subjects"],
            [
                [
                    split,
                    summary["images_by_split"].get(split, 0),
                    summary["components_by_split"].get(split, 0),
                    summary["subjects_by_split"].get(split, 0),
                ]
                for split in SPLIT_ORDER
            ],
        ),
        "",
        "## Binary Label Distribution By Split",
        "",
        markdown_table(
            ["Split", "NonDemented", "Demented"],
            [
                [
                    split,
                    summary["labels_by_split"].get(split, {}).get("NonDemented", 0),
                    summary["labels_by_split"].get(split, {}).get("Demented", 0),
                ]
                for split in SPLIT_ORDER
            ],
        ),
        "",
        "## Leakage Safety Checks",
        "",
    ]
    for key, values in summary["component_overlap"].items():
        lines.append(f"- Component overlap `{key}`: **{len(values)}**")
    for key, values in summary["subject_overlap"].items():
        lines.append(f"- Subject overlap `{key}`: **{len(values)}**")
    lines.extend(["", "## QC Decision", ""])
    if summary["overlap_check_passed"]:
        lines.append("**GO for OASIS-1 baseline training from this frozen split manifest.**")
    else:
        lines.append("**NO-GO: subject/component overlap detected.**")
    SPLIT_AUDIT_MD.parent.mkdir(parents=True, exist_ok=True)
    SPLIT_AUDIT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(
    inventory_rows: list[dict[str, object]],
    extraction_summary: dict[str, object],
    volume_rows: list[dict[str, object]],
    slice_rows: list[dict[str, object]],
    split_summary: dict,
) -> None:
    summary = {
        "created_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source_dataset": SOURCE_DATASET,
        "archives": {
            "count": len(list(TAR_DIR.glob("oasis_cross-sectional_disc*.tar"))),
            "sessions": len(inventory_rows),
            "unique_subjects": len({row["subject_id"] for row in inventory_rows}),
        },
        "extraction": extraction_summary,
        "volume_manifest": {
            "path": str(VOLUME_MANIFEST_CSV),
            "rows": len(volume_rows),
            "label_counts": dict(Counter(row["binary_label"] for row in volume_rows)),
        },
        "slice_manifest": {
            "path": str(SLICE_MANIFEST_CSV),
            "rows": len(slice_rows),
            "label_counts": dict(Counter(row["binary_label"] for row in slice_rows)),
        },
        "split": split_summary,
        "outputs": {
            "inventory_csv": str(INVENTORY_CSV),
            "inventory_audit": str(INVENTORY_AUDIT_MD),
            "manifest_audit": str(MANIFEST_AUDIT_MD),
            "component_csv": str(COMPONENT_CSV),
            "split_csv": str(SPLIT_CSV),
            "split_audit": str(SPLIT_AUDIT_MD),
        },
    }
    SUMMARY_JSON.parent.mkdir(parents=True, exist_ok=True)
    SUMMARY_JSON.write_text(json.dumps(summary, indent=2), encoding="utf-8")


def run_pipeline(args: argparse.Namespace) -> int:
    metadata_rows = read_metadata()
    inventory_rows = build_inventory(metadata_rows)
    write_csv(INVENTORY_CSV, inventory_rows, list(inventory_rows[0].keys()))
    write_inventory_audit(inventory_rows)

    extraction_summary = {"skipped": True}
    if args.extract:
        extraction_summary = extract_selected_volumes(metadata_rows, force=args.force)
        inventory_rows = build_inventory(metadata_rows)
        write_csv(INVENTORY_CSV, inventory_rows, list(inventory_rows[0].keys()))
        write_inventory_audit(inventory_rows)

    volume_rows = build_volume_manifest(metadata_rows)
    slice_rows = []
    if args.slices:
        slice_rows = generate_slices(volume_rows, force=args.force)
    elif SLICE_MANIFEST_CSV.exists():
        with SLICE_MANIFEST_CSV.open(newline="", encoding="utf-8") as handle:
            slice_rows = list(csv.DictReader(handle))

    write_manifest_audit(volume_rows, slice_rows)
    component_rows = build_components(slice_rows) if slice_rows else []
    split_summary = {
        "total_images": 0,
        "overlap_check_passed": False,
        "note": "No slices available for split generation.",
    }
    if component_rows:
        _, split_summary = make_split(component_rows, seed=args.seed)
        write_split_audit(split_summary)

    write_summary(inventory_rows, extraction_summary, volume_rows, slice_rows, split_summary)
    print(f"Wrote inventory: {INVENTORY_CSV}")
    print(f"Wrote volume manifest: {VOLUME_MANIFEST_CSV}")
    if slice_rows:
        print(f"Wrote slice manifest: {SLICE_MANIFEST_CSV}")
        print(f"Wrote split manifest: {SPLIT_CSV}")
    print(f"Wrote summary: {SUMMARY_JSON}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--extract", action="store_true", help="Selectively extract processed masked T88 volumes.")
    parser.add_argument("--slices", action="store_true", help="Generate central coronal slice PNGs.")
    parser.add_argument("--force", action="store_true", help="Overwrite selected volumes and derived slices.")
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()
    return run_pipeline(args)


if __name__ == "__main__":
    raise SystemExit(main())
